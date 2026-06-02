"""Inspeciona a planilha oficial que o Caio usa hoje."""
from __future__ import annotations

import openpyxl
from pathlib import Path


def main():
    f = Path(r"C:\Users\caio.ferreira\Aura Minerals\Borborema - Planta - Beneficiamento - Documentos\Beneficiamento\Operacao_Planta\Result. Laboratório\Resultados Laboratorio_Oficial NOVA.xlsx")
    wb = openpyxl.load_workbook(f, data_only=True)
    print(f"=== {f.name} ===")
    print(f"Sheets ({len(wb.sheetnames)}): {wb.sheetnames}")
    print()

    for sname in wb.sheetnames:
        ws = wb[sname]
        print("=" * 90)
        print(f"--- Sheet: {sname!r} ({ws.max_row} linhas x {ws.max_column} colunas) ---")
        print("=" * 90)

        # Merges
        if ws.merged_cells.ranges:
            print(f"Merges ({len(ws.merged_cells.ranges)}):")
            for m in list(ws.merged_cells.ranges)[:30]:
                top_left = ws.cell(m.min_row, m.min_col).value
                if top_left:
                    print(f"  {m.coord:18s} = {str(top_left)[:60]}")
            print()

        # Primeiras 20 linhas
        max_rows = min(ws.max_row, 20)
        max_cols = min(ws.max_column, 30)
        for r in range(1, max_rows + 1):
            cells = []
            for c in range(1, max_cols + 1):
                v = ws.cell(r, c).value
                if v is None:
                    cells.append("")
                else:
                    s = str(v).strip()
                    cells.append(s[:25] + "..." if len(s) > 25 else s)
            if any(c for c in cells):
                non_empty = [(i + 1, c) for i, c in enumerate(cells) if c]
                parts = [f"{openpyxl.utils.get_column_letter(i)}={c!r}" for i, c in non_empty]
                print(f"  R{r:03d}: {' | '.join(parts)}")
        if ws.max_row > max_rows:
            print(f"  ... (+ {ws.max_row - max_rows} linhas)")
        print()


if __name__ == "__main__":
    main()
