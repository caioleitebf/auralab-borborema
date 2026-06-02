"""Inspeciona o arquivo API (1).xlsx — versao detalhada.

Mostra valores e formulas, e tenta identificar regioes/blocos de tabela
por aglomeracao de celulas preenchidas.
"""
from __future__ import annotations

import openpyxl
from pathlib import Path


def main():
    f = Path(r"C:\Users\caio.ferreira\Downloads\API (1).xlsx")
    # Carrega sem data_only para ver formulas
    wb_form = openpyxl.load_workbook(f, data_only=False)
    wb_val = openpyxl.load_workbook(f, data_only=True)

    for sname in wb_form.sheetnames:
        ws_f = wb_form[sname]
        ws_v = wb_val[sname]
        print(f"=== Sheet: {sname!r} ({ws_f.max_row} linhas x {ws_f.max_column} colunas) ===")
        print()

        # Detecta regioes/blocos por linhas continuas com conteudo
        bloco_atual = []
        ultimo_r_com_conteudo = 0
        for r in range(1, ws_f.max_row + 1):
            tem_conteudo = False
            for c in range(1, ws_f.max_column + 1):
                if ws_f.cell(r, c).value is not None:
                    tem_conteudo = True
                    break
            if tem_conteudo:
                if r - ultimo_r_com_conteudo > 3 and bloco_atual:
                    print(f"\n--- BLOCO (linhas {bloco_atual[0]} a {bloco_atual[-1]}) ---")
                    bloco_atual = []
                bloco_atual.append(r)
                ultimo_r_com_conteudo = r
        if bloco_atual:
            print(f"\n--- BLOCO (linhas {bloco_atual[0]} a {bloco_atual[-1]}) ---")

        # Agora mostra tudo na sheet com TODAS celulas (mesmo vazias)
        print()
        print("=== DUMP COMPLETO (col A-X, todas linhas com algo) ===")
        for r in range(1, ws_f.max_row + 1):
            cells_f = [ws_f.cell(r, c).value for c in range(1, ws_f.max_column + 1)]
            cells_v = [ws_v.cell(r, c).value for c in range(1, ws_f.max_column + 1)]
            if all(c is None for c in cells_f):
                continue
            line = f"R{r:03d}: "
            for i, (cf, cv) in enumerate(zip(cells_f, cells_v)):
                if cf is None and cv is None:
                    continue
                col_letter = openpyxl.utils.get_column_letter(i + 1)
                if cf == cv or cv is None:
                    s = str(cf)[:30]
                else:
                    s = f"{str(cf)[:20]}={str(cv)[:15]}"
                line += f"{col_letter}={s!r} "
            print(line)


if __name__ == "__main__":
    main()
