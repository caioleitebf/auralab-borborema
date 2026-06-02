"""Exportacao de dados para Excel multi-aba.

Cada aba do Excel = um processo. Usa openpyxl para formatacao com cabecalho
Aura (azul-marinho + coral).
"""
from __future__ import annotations

import io
from datetime import date

import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from ..config import Cores
from . import canonical


# Cores em formato openpyxl (sem #)
_FILL_BLUE = PatternFill(start_color="2D3D70", end_color="2D3D70", fill_type="solid")
_FILL_BLUE_MED = PatternFill(start_color="3C4788", end_color="3C4788", fill_type="solid")
_FILL_CORAL = PatternFill(start_color="F4614D", end_color="F4614D", fill_type="solid")
_FILL_CORAL_MED = PatternFill(start_color="FF9477", end_color="FF9477", fill_type="solid")
_FILL_CORAL_BG = PatternFill(start_color="FFCFC3", end_color="FFCFC3", fill_type="solid")
_FILL_ALT = PatternFill(start_color="F7F9FC", end_color="F7F9FC", fill_type="solid")

_FONT_HEADER = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
_FONT_BODY = Font(name="Segoe UI", size=10, color="2D3D70")

_ALIGN_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
_BORDER_THIN = Border(
    left=Side(border_style="thin", color="E5E7EF"),
    right=Side(border_style="thin", color="E5E7EF"),
    top=Side(border_style="thin", color="E5E7EF"),
    bottom=Side(border_style="thin", color="E5E7EF"),
)


def _style_sheet(ws, n_header_rows: int = 1):
    """Aplica estilo Aura nas N primeiras linhas do header + corpo."""
    max_col = ws.max_column
    max_row = ws.max_row

    # Identifica colunas que contem 'DATA' no cabecalho (para formato BR)
    cols_data = set()
    for c in range(1, max_col + 1):
        for r in range(1, n_header_rows + 1):
            v = ws.cell(row=r, column=c).value
            if v and "DATA" in str(v).upper():
                cols_data.add(c)
                break

    # Header
    for r in range(1, n_header_rows + 1):
        for c in range(1, max_col + 1):
            cell = ws.cell(row=r, column=c)
            cell.fill = _FILL_BLUE
            cell.font = _FONT_HEADER
            cell.alignment = _ALIGN_CENTER
            cell.border = _BORDER_THIN
        ws.row_dimensions[r].height = 22

    # Body
    for r in range(n_header_rows + 1, max_row + 1):
        for c in range(1, max_col + 1):
            cell = ws.cell(row=r, column=c)
            cell.font = _FONT_BODY
            cell.alignment = _ALIGN_CENTER
            cell.border = _BORDER_THIN
            if r % 2 == 0:
                cell.fill = _FILL_ALT
            # Datas em formato BR
            if c in cols_data and cell.value is not None:
                cell.number_format = "DD/MM/YYYY"

    # Larguras automaticas (heuristica)
    for c in range(1, max_col + 1):
        col_letter = get_column_letter(c)
        max_len = 8
        for r in range(1, min(max_row + 1, 50)):
            v = ws.cell(row=r, column=c).value
            if v is not None:
                max_len = max(max_len, len(str(v)) + 2)
        ws.column_dimensions[col_letter].width = min(max_len + 2, 22)

    # Freeze panes (cabecalho)
    ws.freeze_panes = ws.cell(row=n_header_rows + 1, column=1)


def gerar_excel_consolidado(di: date, df: date) -> bytes:
    """Gera Excel multi-aba com todos os processos do periodo.

    Returns: bytes do .xlsx pronto para download.
    """
    abas = [
        ("Lixiviacao", canonical.lixiviacao_wide(di, df)),
        ("Tanques (1045)", canonical.tanques_wide(di, df)),
        ("Acacia", canonical.acacia_wide(di, df)),
        ("Eluicao", canonical.eluicao_wide(di, df)),
        ("Eletrolise CE0001", canonical.eletrolise_wide(di, df, "CE0001")),
        ("Eletrolise CE0002", canonical.eletrolise_wide(di, df, "CE0002")),
        ("Agua de Processo", canonical.agua_processo_wide(di, df)),
        ("Detox", canonical.detox_wide(di, df)),
        ("Bullion", canonical.bullion_wide(di, df)),
    ]

    buffer = io.BytesIO()
    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        # Aba de resumo (capa)
        capa = pd.DataFrame({
            "Processo": [a[0] for a in abas],
            "Registros": [len(a[1]) for a in abas],
        })
        capa.to_excel(writer, sheet_name="Resumo", index=False, startrow=1)
        ws = writer.sheets["Resumo"]
        ws.cell(row=1, column=1, value=f"AuraLab Borborema — Exportação Consolidada")
        ws.cell(row=1, column=2, value=f"Período: {di:%d/%m/%Y} a {df:%d/%m/%Y}")
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=2)
        ws.cell(row=1, column=1).fill = _FILL_CORAL
        ws.cell(row=1, column=1).font = _FONT_HEADER
        ws.cell(row=1, column=1).alignment = _ALIGN_CENTER

        for nome, df_aba in abas:
            if df_aba.empty:
                # Cria aba vazia com mensagem
                pd.DataFrame({"Sem dados no periodo": []}).to_excel(
                    writer, sheet_name=nome[:31], index=False)
                continue
            # Achata MultiIndex se houver
            df_export = df_aba.copy()
            if isinstance(df_export.columns, pd.MultiIndex):
                df_export.columns = [" ".join(filter(None, c)).strip() for c in df_export.columns]
            df_export.to_excel(writer, sheet_name=nome[:31], index=False)

        # Aplica estilo Aura em todas as abas
        for nome, df_aba in abas:
            sheet_name = nome[:31]
            if sheet_name in writer.sheets:
                _style_sheet(writer.sheets[sheet_name])

        # Estilo da capa
        _style_sheet(writer.sheets["Resumo"], n_header_rows=2)

    return buffer.getvalue()


def gerar_excel_aba(nome_aba: str, df: pd.DataFrame) -> bytes:
    """Gera Excel de uma unica aba (para download local)."""
    buffer = io.BytesIO()
    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        if df.empty:
            pd.DataFrame({"Sem dados": []}).to_excel(writer, sheet_name=nome_aba[:31], index=False)
        else:
            df_export = df.copy()
            if isinstance(df_export.columns, pd.MultiIndex):
                df_export.columns = [" ".join(filter(None, c)).strip() for c in df_export.columns]
            df_export.to_excel(writer, sheet_name=nome_aba[:31], index=False)
        sheet_name = nome_aba[:31]
        if sheet_name in writer.sheets:
            _style_sheet(writer.sheets[sheet_name])
    return buffer.getvalue()
